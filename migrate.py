"""
Skrip migrasi data dari Excel ke Supabase.
Jalankan: python migrate.py

Pastikan sudah install:
  pip install openpyxl supabase

Isi SUPABASE_URL dan SUPABASE_SERVICE_KEY di bawah.
Gunakan Service Role Key (bukan anon key) agar bisa bypass RLS.
"""

import openpyxl
from supabase import create_client

# ===================================================
# KONFIGURASI - ISI SESUAI PROJECT SUPABASE ANDA
# ===================================================
SUPABASE_URL = "https://xmgnsalzsngxawquqdrw.supabase.co"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InhtZ25zYWx6c25neGF3cXVxZHJ3Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3ODAyODA1MywiZXhwIjoyMDkzNjA0MDUzfQ.HsxM7R3ui6usIL0UX4ILbU5JQIM-Q8nnrYsRR7T7rfo"  # Service Role Key dari Settings > API
EXCEL_PATH = r"c:\iuran_air\daftar pengguna air.xlsx"
# ===================================================

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# Mapping: kolom Excel (index 2-13) -> bulan (1-12)
# Urutan kolom: Okt, Nov, Des, Jan, Feb, Mar, Apr, Mei, Jun, Jul, Agu, Sep
KOLOM_BULAN = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]

def get_tahun_id(label):
    res = sb.table("tahun_iuran").select("id").eq("label", label).single().execute()
    return res.data["id"]

def migrate_sheet(sheet, tahun_label):
    print(f"\n=== Migrasi sheet: {sheet.title} (Tahun {tahun_label}) ===")
    tahun_id = get_tahun_id(tahun_label)

    anggota_list = []
    pembayaran_list = []

    for row in sheet.iter_rows(min_row=4, values_only=True):
        no = row[0]
        nama = row[1]
        if not no or not nama or not isinstance(no, int):
            continue

        nama = str(nama).strip()
        if not nama:
            continue

        # Tentukan nominal_iuran dari baris (ambil nilai pertama yang ada)
        nominal_iuran = 20000
        for col_i, bulan in enumerate(KOLOM_BULAN):
            val = row[2 + col_i]
            if isinstance(val, (int, float)) and val > 0:
                nominal_iuran = int(val)
                break

        # Bulatkan ke kelipatan 20000
        nominal_iuran = max(20000, (nominal_iuran // 20000) * 20000)

        anggota_list.append({
            "nomor_urut": no,
            "nama": nama,
            "nominal_iuran": nominal_iuran,
            "is_active": True,
        })

    # Upsert pengguna
    for anggota in anggota_list:
        res = sb.table("pengguna").upsert(anggota, on_conflict="nomor_urut").execute()

    # Ambil ulang pengguna yang sudah tersimpan (untuk dapat id)
    pengguna_db = sb.table("pengguna").select("id, nomor_urut").execute().data
    no_to_id = {p["nomor_urut"]: p["id"] for p in pengguna_db}

    print(f"  {len(anggota_list)} pengguna diproses.")

    # Proses pembayaran
    for row in sheet.iter_rows(min_row=4, values_only=True):
        no = row[0]
        if not no or not isinstance(no, int):
            continue
        pengguna_id = no_to_id.get(no)
        if not pengguna_id:
            continue

        for col_i, bulan in enumerate(KOLOM_BULAN):
            val = row[2 + col_i]
            if isinstance(val, (int, float)) and val > 0:
                nominal = (int(val) // 20000) * 20000
                if nominal <= 0:
                    nominal = 20000
                pembayaran_list.append({
                    "pengguna_id": pengguna_id,
                    "tahun_iuran_id": tahun_id,
                    "bulan": bulan,
                    "nominal": nominal,
                    "tanggal_bayar": f"{tahun_label}-01-01",  # tanggal estimasi
                })

    if pembayaran_list:
        # Upsert berdasarkan unique constraint (pengguna_id, tahun_iuran_id, bulan)
        sb.table("pembayaran").upsert(
            pembayaran_list,
            on_conflict="pengguna_id,tahun_iuran_id,bulan"
        ).execute()
        print(f"  {len(pembayaran_list)} record pembayaran dimigrasi.")

def migrate_pemasukan(sheet):
    print("\n=== Migrasi Pemasukan ===")
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] and row[3] and isinstance(row[3], (int, float)):
            tanggal = row[1].strftime("%Y-%m-%d") if hasattr(row[1], 'strftime') else "2024-12-01"
            rows.append({
                "tanggal": tanggal,
                "keterangan": str(row[2]).strip(),
                "nominal": int(row[3]),
            })
    if rows:
        sb.table("pemasukan").insert(rows).execute()
        print(f"  {len(rows)} pemasukan dimigrasi.")

def migrate_pengeluaran(sheet):
    print("\n=== Migrasi Pengeluaran ===")
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        no = row[0]
        tanggal_raw = row[1]
        keterangan = row[2]
        nominal = row[3]
        if not isinstance(no, int) or not keterangan or not isinstance(nominal, (int, float)):
            continue
        try:
            tanggal = tanggal_raw.strftime("%Y-%m-%d") if hasattr(tanggal_raw, 'strftime') else "2024-11-25"
        except Exception:
            tanggal = "2024-11-25"
        rows.append({
            "tanggal": tanggal,
            "keterangan": str(keterangan).strip(),
            "nominal": int(nominal),
        })
    if rows:
        sb.table("pengeluaran").insert(rows).execute()
        print(f"  {len(rows)} pengeluaran dimigrasi.")

if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")

    print("Membuka file Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    migrate_sheet(wb["iuran 2024"], "2024")
    migrate_sheet(wb["iuran oktober 2025"], "2025")
    migrate_pemasukan(wb["pemasukan"])
    migrate_pengeluaran(wb["pengeluaran"])

    print("\n✓ Migrasi selesai!")
    print("Buka Supabase Dashboard untuk verifikasi data.")
