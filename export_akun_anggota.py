"""
Export daftar akun anggota (email + password default) ke file CSV dan Excel.
Jalankan: python export_akun_anggota.py
"""

import re
import csv
import os
from supabase import create_client

SUPABASE_URL = "https://xmgnsalzsngxawquqdrw.supabase.co"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InhtZ25zYWx6c25neGF3cXVxZHJ3Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3ODAyODA1MywiZXhwIjoyMDkzNjA0MDUzfQ.HsxM7R3ui6usIL0UX4ILbU5JQIM-Q8nnrYsRR7T7rfo"
PASSWORD_DEFAULT = "lebaksiuh"
OUTPUT_CSV   = "daftar_akun_anggota.csv"
OUTPUT_EXCEL = "daftar_akun_anggota.xlsx"


def nama_ke_email(nama):
    nama = nama.strip().lower()
    nama = nama.replace("'", "")
    if '/' in nama:
        parts = nama.split('/')
        parts = [''.join(p.split()) for p in parts]
        local = '.'.join(p for p in parts if p)
    else:
        local = ''.join(nama.split())
        local = re.sub(r'\.{2,}', '.', local)
    local = local.strip('.')
    return f"{local}@iuranair.com"


sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
pengguna_list = sb.table("pengguna").select("nomor_urut, nama").order("nomor_urut").execute().data

rows = []
for p in pengguna_list:
    rows.append({
        "No": p["nomor_urut"],
        "Nama": p["nama"].strip(),
        "Email": nama_ke_email(p["nama"]),
        "Password": PASSWORD_DEFAULT,
        "URL Aplikasi": "https://lebaksiuh.github.io/iuran_air/",
    })

# === Export CSV ===
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=["No", "Nama", "Email", "Password", "URL Aplikasi"])
    writer.writeheader()
    writer.writerows(rows)
print(f"CSV tersimpan: {OUTPUT_CSV}")

# === Export Excel (jika openpyxl tersedia) ===
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Akun"

    # Header
    headers = ["No", "Nama", "Email / Username", "Password", "URL Aplikasi"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20

    # Data
    for row_idx, r in enumerate(rows, 2):
        vals = [r["No"], r["Nama"], r["Email"], r["Password"], r["URL Aplikasi"]]
        fill = PatternFill("solid", fgColor="F0F7FF") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="center")
            cell.fill = fill
            cell.border = border

    # Lebar kolom
    col_widths = [6, 30, 38, 15, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_EXCEL)
    print(f"Excel tersimpan: {OUTPUT_EXCEL}")

except ImportError:
    print("Info: openpyxl tidak terinstall, hanya CSV yang dibuat.")
    print("      Install dengan: pip install openpyxl")

print(f"\nTotal: {len(rows)} akun")
print(f"Password default: {PASSWORD_DEFAULT}")
